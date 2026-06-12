"""종목 스크리너 — PBR × ROE 4사분면."""

import io
import sys
from pathlib import Path
from datetime import date

import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, str(Path(__file__).parent.parent))

from utils.styles import inject_css, COLOR
from utils.charts import apply_layout, bubble_chart, heatmap_chart
from utils.valuation import get_investment_signal, get_badge_label, get_quadrant
from data.cache import cached_all_stocks
from data.fetcher import is_etf_etn

st.set_page_config(page_title="종목 스크리너", page_icon="🔍", layout="wide")
inject_css()

# ──────────────────────────────────────────────
# 섹터 매핑 (Naver 업종 기반 근사 분류)
# ──────────────────────────────────────────────
SECTOR_MAP: dict[str, str] = {}  # ticker → sector (추후 확장)

ALL_SECTORS = [
    "반도체/전자", "자동차", "조선/중공업", "건설/부동산",
    "금융/보험", "화학/소재", "바이오/헬스케어", "IT서비스",
    "방산/항공", "에너지/유틸리티", "소비재/유통", "기타",
]

# ──────────────────────────────────────────────
# 데이터 로드
# ──────────────────────────────────────────────
@st.cache_data(ttl=3600, show_spinner=False)
def load_data(market: str) -> pd.DataFrame:
    return cached_all_stocks(market)


# ──────────────────────────────────────────────
# 사이드바 필터
# ──────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🔍 스크리닝 조건")

    market = st.selectbox("시장", ["전체", "KOSPI", "KOSDAQ"])
    pbr_max = st.slider("PBR 최대", 0.1, 5.0, 1.5, step=0.1)
    roe_min = st.slider("ROE 최소 (%)", -20, 50, 15)
    per_max = st.slider("PER 최대", 0, 100, 30)
    cap_filter = st.selectbox(
        "시가총액 구간",
        ["전체", "대형주", "중형주", "소형주"],
        help="KOSPI: 대형(1~100위)/중형(101~300위)/소형(301위↓) | KOSDAQ: 대형(1~100위)/중형(101~400위)/소형(401위↓)",
    )

    st.divider()
    show_all = st.checkbox("필터 해제 (전종목 표시)", value=False)
    include_etf = st.checkbox("ETF/ETN 포함", value=False)
    st.caption(f"마지막 업데이트: {date.today().strftime('%Y-%m-%d')}")

# ──────────────────────────────────────────────
# 메인 영역
# ──────────────────────────────────────────────
st.title("🔍 PBR × ROE 종목 스크리너")
st.caption("저PBR + 고ROE = 저평가 우량주 발굴 | PBR = PER × ROE")

# 데이터 로드
market_arg = "ALL" if market == "전체" else market
with st.spinner(f"{market} 데이터 수집 중..."):
    raw_df = load_data(market_arg)

if raw_df.empty:
    st.error("데이터를 불러올 수 없습니다. 잠시 후 다시 시도해 주세요.")
    st.stop()

# ETF/ETN 제외 (is_etf 컬럼 없는 구버전 캐시 fallback 포함)
if "is_etf" not in raw_df.columns:
    raw_df = raw_df.copy()
    raw_df["is_etf"] = raw_df["name"].apply(is_etf_etn)

if not include_etf:
    raw_df = raw_df[~raw_df["is_etf"]]

# 시장별 시가총액 순위 계산 (전체 데이터 기준)
raw_df = raw_df.copy()
for _mkt in ["KOSPI", "KOSDAQ"]:
    _mask = raw_df["market"] == _mkt
    raw_df.loc[_mask, "cap_rank"] = (
        raw_df.loc[_mask, "marcap"]
        .rank(ascending=False, method="min")
        .astype("Int64")
    )

# PBR > 0 필터 (음수 자기자본 제외)
df = raw_df[raw_df["pbr"] > 0].copy() if not show_all else raw_df.copy()


def _quad_label(pbr: float, roe: float) -> str:
    """슬라이더 임계값 기준 4사분면 레이블."""
    if pbr <= pbr_max:
        return "저평가 우량" if roe >= roe_min else "가치함정"
    return "고평가 우량" if roe >= roe_min else "고평가 위험"

# 시가총액 순위 기반 필터
if cap_filter != "전체":
    _kp = df["market"] == "KOSPI"
    _kd = df["market"] == "KOSDAQ"
    if cap_filter == "대형주":
        df = df[(_kp & (df["cap_rank"] <= 100)) | (_kd & (df["cap_rank"] <= 100))]
    elif cap_filter == "중형주":
        df = df[
            (_kp & (df["cap_rank"] >= 101) & (df["cap_rank"] <= 300)) |
            (_kd & (df["cap_rank"] >= 101) & (df["cap_rank"] <= 400))
        ]
    elif cap_filter == "소형주":
        df = df[(_kp & (df["cap_rank"] >= 301)) | (_kd & (df["cap_rank"] >= 401))]

# 버블차트용 전체 df 저장 (PBR/ROE/PER 필터 전, 색상 구분용)
chart_df = df[df["pbr"] > 0].dropna(subset=["pbr", "roe"]).copy()
chart_df["판단"] = chart_df.apply(lambda r: _quad_label(r["pbr"], r["roe"]), axis=1)

# PBR/ROE/PER 필터 (show_all이 아닐 때) — 테이블·지표용
if not show_all:
    df = df[
        (df["pbr"].isna() | (df["pbr"] <= pbr_max)) &
        (df["roe"].isna() | (df["roe"] >= roe_min)) &
        (df["per"].isna() | (df["per"] <= per_max))
    ]

df = df.dropna(subset=["pbr", "roe"]).copy()

# 투자 판단 추가
df["판단"] = df.apply(lambda r: _quad_label(r["pbr"], r["roe"]), axis=1)
df = df.sort_values("pbr", ascending=True).reset_index(drop=True)

# 요약 메트릭
col1, col2, col3, col4 = st.columns(4)
with col1:
    st.metric("검색 결과", f"{len(df)}종목")
with col2:
    n_value = len(df[(df.pbr <= 1.0) & (df.roe >= 15)])
    st.metric("저평가 우량", f"{n_value}종목",
              help="PBR≤1.0 & ROE≥15%")
with col3:
    avg_pbr = df["pbr"].mean()
    st.metric("평균 PBR", f"{avg_pbr:.2f}x")
with col4:
    avg_roe = df["roe"].mean()
    st.metric("평균 ROE", f"{avg_roe:.1f}%")

st.divider()

# ──────────────────────────────────────────────
# PBR × ROE 버블차트
# ──────────────────────────────────────────────
tab1, tab2, tab3 = st.tabs(["📊 버블차트 (4사분면)", "🗺️ 섹터 히트맵", "📋 종목 테이블"])

with tab1:
    st.subheader("PBR × ROE 4사분면 버블차트")
    st.caption("버블 크기: 시가총액 | 색상: 투자 판단 영역")

    # 색상 매핑
    quadrant_colors = {
        "low_pbr_high_roe": COLOR["primary"],    # 파란 (목표)
        "high_pbr_high_roe": COLOR["text_muted"], # 회색
        "low_pbr_low_roe": COLOR["warning"],      # 노란 (주의)
        "high_pbr_low_roe": COLOR["negative"],    # 빨간
    }

    fig = go.Figure()

    # chart_df 기준으로 전체 종목 표시 (슬라이더는 구분선만 이동)
    _max_cap = chart_df["marcap"].max() or 1
    for q_key, q_color in quadrant_colors.items():
        mask = chart_df.apply(
            lambda r: get_quadrant(r["pbr"], r["roe"],
                                   pbr_threshold=pbr_max,
                                   roe_threshold=float(roe_min)) == q_key,
            axis=1,
        )
        sub = chart_df[mask]
        if sub.empty:
            continue

        q_labels = {
            "low_pbr_high_roe":  "저평가 우량 — 적극 매수 검토",
            "high_pbr_high_roe": "고평가 우량 — 진입 시점 주의",
            "low_pbr_low_roe":   "가치함정 — 실적 개선 확인",
            "high_pbr_low_roe":  "고평가 위험 — 회피",
        }

        # 버블 크기 정규화 (전체 기준으로 일관되게)
        size_vals = sub["marcap"].fillna(100)
        size_norm = (size_vals / _max_cap * 24 + 4).clip(4, 28)

        fig.add_trace(go.Scatter(
            x=sub["roe"], y=sub["pbr"],
            mode="markers",
            name=q_labels[q_key],
            marker=dict(
                size=size_norm,
                color=q_color,
                opacity=0.60,
                line=dict(width=0.5, color="white"),
            ),
            text=sub["name"],
            customdata=sub[["ticker", "per", "marcap"]].values,
            hovertemplate=(
                "<b>%{text}</b> (%{customdata[0]})<br>"
                "ROE: %{x:.1f}%<br>"
                "PBR: %{y:.2f}x<br>"
                "PER: %{customdata[1]:.1f}x<br>"
                "시가총액: %{customdata[2]:,.0f}억<br>"
                "<extra></extra>"
            ),
        ))

    # 기준선 (슬라이더 조건값 기준 — 동적)
    fig.add_vline(x=float(roe_min), line_dash="dash", line_color=COLOR["primary"],
                  line_width=1.2,
                  annotation_text=f"ROE {roe_min}%", annotation_position="top right",
                  annotation_font_color=COLOR["primary"])
    fig.add_hline(y=pbr_max, line_dash="dash", line_color=COLOR["primary"],
                  line_width=1.2,
                  annotation_text=f"PBR {pbr_max:.1f}x", annotation_position="top right",
                  annotation_font_color=COLOR["primary"])

    apply_layout(fig, "")
    fig.update_layout(
        xaxis_title="ROE (%)",
        yaxis_title="PBR (배)",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        height=550,
    )
    # 축 범위 제한 (전체 chart_df 기준 극단값 클리핑)
    roe_clip = chart_df["roe"].clip(-20, 80)
    pbr_clip = chart_df["pbr"].clip(0, 8)
    fig.update_xaxes(range=[roe_clip.quantile(0.01) - 5, roe_clip.quantile(0.99) + 5])
    fig.update_yaxes(range=[0, min(pbr_clip.quantile(0.99) + 1, 8)])

    chart_event = st.plotly_chart(fig, on_select="rerun", key="bubble_chart", use_container_width=True)

    # 버블 클릭 → 기업 분석 연동
    if chart_event and chart_event.selection and chart_event.selection.points:
        pt = chart_event.selection.points[0]
        raw_cd = pt.get("customdata") or []
        sel_ticker = str(raw_cd[0]).zfill(6) if raw_cd else None
        if sel_ticker:
            matched = df[df["ticker"] == sel_ticker]
            if not matched.empty:
                sel_name = matched.iloc[0]["name"]
                c_info, c_btn = st.columns([3, 1])
                with c_info:
                    st.info(f"**{sel_name}** ({sel_ticker}) 선택됨")
                with c_btn:
                    if st.button("기업 분석 →", type="primary", key="chart_goto"):
                        st.session_state.selected_ticker = sel_ticker
                        st.session_state.selected_name = sel_name
                        st.switch_page("pages/03_company.py")

    with st.expander("4사분면 해석"):
        st.caption("차트 구조: X축=ROE(낮음→높음), Y축=PBR(낮음→높음) | 기준선: 슬라이더 설정값")
        st.markdown(f"""
        | 위치 (차트 기준) | PBR | ROE | 해석 | 전략 |
        |--------|-----|-----|------|------|
        | 🔵 **우하단** | ≤{pbr_max:.1f}x | ≥{roe_min}% | **저평가 우량** — 핵심 목표 영역 | 적극 매수 검토 |
        | ⚪ **우상단** | >{pbr_max:.1f}x | ≥{roe_min}% | 고평가 우량 — 이미 시장 반영 | 진입 시점 주의 |
        | 🟡 **좌하단** | ≤{pbr_max:.1f}x | <{roe_min}% | **가치함정** — 싸지만 수익성 낮음 | 실적 개선 확인 필수 |
        | 🔴 **좌상단** | >{pbr_max:.1f}x | <{roe_min}% | 고평가 위험 — 최악의 조합 | 회피 |

        > X축=ROE, Y축=PBR이므로 **저PBR+고ROE(목표 영역)는 우하단**에 위치합니다.
        """)

with tab2:
    st.subheader("시장 전체 분포 현황")

    # 판단별 분포 바차트 (전체 종목 기준 — 4사분면 모두 표시)
    signal_counts = chart_df["판단"].value_counts().reset_index()
    signal_counts.columns = ["판단", "종목수"]

    signal_color_map = {
        "저평가 우량": COLOR["primary"],
        "고평가 우량": COLOR["text_muted"],
        "가치함정":    COLOR["warning"],
        "고평가 위험": COLOR["negative"],
    }

    fig_bar = go.Figure()
    for _, row in signal_counts.iterrows():
        fig_bar.add_trace(go.Bar(
            x=[row["판단"]], y=[row["종목수"]],
            marker_color=signal_color_map.get(row["판단"], COLOR["text_muted"]),
            name=row["판단"],
            showlegend=False,
            text=[row["종목수"]],
            textposition="outside",
        ))

    apply_layout(fig_bar, "")
    fig_bar.update_layout(
        xaxis_title="투자 판단",
        yaxis_title="종목 수",
        height=350,
    )
    st.plotly_chart(fig_bar, use_container_width=True)

    # PBR 분포 히스토그램
    st.caption("PBR 분포")
    fig_hist = go.Figure()
    pbr_clipped = df["pbr"].clip(0, 5)
    fig_hist.add_trace(go.Histogram(
        x=pbr_clipped, nbinsx=50,
        marker_color=COLOR["primary"], opacity=0.7,
    ))
    fig_hist.add_vline(x=1.0, line_dash="dash", line_color=COLOR["negative"],
                       annotation_text="PBR 1.0x")
    apply_layout(fig_hist, "")
    fig_hist.update_layout(xaxis_title="PBR", yaxis_title="종목 수", height=280)
    st.plotly_chart(fig_hist, use_container_width=True)

with tab3:
    st.subheader("종목 목록")

    # ── 테이블 내 추가 필터 ──
    with st.expander("🔧 테이블 내 필터", expanded=False):
        tcol1, tcol2, tcol3 = st.columns(3)
        with tcol1:
            name_filter = st.text_input("종목명 검색", placeholder="예: 삼성, 현대")
        with tcol2:
            signal_options = ["전체", "저평가 우량", "고평가 우량", "가치함정", "고평가 위험"]
            signal_filter = st.selectbox("투자 판단", signal_options)
        with tcol3:
            sort_by = st.selectbox("정렬 기준", ["PBR 오름차순", "ROE 내림차순", "시가총액 내림차순"])

        tcol4, tcol5 = st.columns(2)
        with tcol4:
            pbr_range = st.slider("PBR 범위", 0.0, 10.0, (0.0, float(chart_df["pbr"].max() if not chart_df.empty else 10.0)),
                                   step=0.1, key="tbl_pbr")
        with tcol5:
            _roe_min = float(chart_df["roe"].min() if not chart_df.empty else -50)
            _roe_max = float(chart_df["roe"].max() if not chart_df.empty else 100)
            roe_range = st.slider("ROE(%) 범위", _roe_min, _roe_max,
                                  (_roe_min, _roe_max), step=1.0, key="tbl_roe")

    # 테이블은 사이드바 PBR/ROE/PER 필터와 무관하게 전체 chart_df 기준
    tbl_df = chart_df.copy()
    if name_filter:
        tbl_df = tbl_df[tbl_df["name"].str.contains(name_filter, na=False) |
                        tbl_df["ticker"].str.contains(name_filter, na=False)]
    if signal_filter != "전체":
        tbl_df = tbl_df[tbl_df["판단"] == signal_filter]
    tbl_df = tbl_df[
        (tbl_df["pbr"] >= pbr_range[0]) & (tbl_df["pbr"] <= pbr_range[1]) &
        (tbl_df["roe"] >= roe_range[0]) & (tbl_df["roe"] <= roe_range[1])
    ]

    sort_map = {
        "PBR 오름차순": ("pbr", True),
        "ROE 내림차순": ("roe", False),
        "시가총액 내림차순": ("marcap", False),
    }
    sort_col, sort_asc = sort_map[sort_by]
    tbl_df = tbl_df.sort_values(sort_col, ascending=sort_asc).reset_index(drop=True)

    st.caption(f"필터 결과: {len(tbl_df)}종목  |  행을 클릭하면 기업 분석 페이지로 이동합니다.")

    # 표시용 DataFrame
    display_df = tbl_df[["ticker", "name", "market", "pbr", "roe", "per", "eps", "marcap", "판단"]].copy()
    display_df.columns = ["코드", "종목명", "시장", "PBR", "ROE(%)", "PER", "EPS(원)", "시가총액(억)", "판단"]
    display_df["PBR"] = display_df["PBR"].round(2)
    display_df["ROE(%)"] = display_df["ROE(%)"].round(2)
    display_df["PER"] = display_df["PER"].round(2)

    _badge_style = {
        "저평가 우량": "background-color: #2F64E3; color: white",
        "고평가 우량": "background-color: #6b7280; color: white",
        "가치함정":    "background-color: #f59e0b; color: white",
        "고평가 위험": "background-color: #ef4444; color: white",
    }
    styled = display_df.style.map(lambda v: _badge_style.get(v, ""), subset=["판단"])

    table_event = st.dataframe(
        styled,
        use_container_width=True,
        hide_index=True,
        selection_mode="single-row",
        on_select="rerun",
        key="stock_table",
    )

    if table_event.selection.rows:
        row_idx = table_event.selection.rows[0]
        sel_row = tbl_df.iloc[row_idx]
        st.session_state.selected_ticker = sel_row["ticker"]
        st.session_state.selected_name = sel_row["name"]
        st.switch_page("pages/03_company.py")

    st.divider()

    # ── XLSX 내보내기 ──
    def create_screening_excel(data: pd.DataFrame) -> bytes:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "전체 종목"

        header_fill = PatternFill(start_color="2F64E3", end_color="2F64E3", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        cols = ["코드", "종목명", "시장", "PBR", "ROE(%)", "PER", "EPS(원)", "시가총액(억)", "판단"]
        for j, col in enumerate(cols, 1):
            cell = ws1.cell(row=1, column=j, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for i, _ in data.iterrows():
            for j, col in enumerate(cols, 1):
                raw_col = {
                    "코드": "ticker", "종목명": "name", "시장": "market",
                    "PBR": "pbr", "ROE(%)": "roe", "PER": "per",
                    "EPS(원)": "eps", "시가총액(억)": "marcap",
                    "판단": "판단_raw",
                }.get(col, col)
                ws1.cell(row=i + 2, column=j, value=data.at[i, raw_col] if raw_col in data.columns else "")

        # Sheet2: 저평가 우량 필터
        ws2 = wb.create_sheet("저평가 우량")
        value_data = df[(df.pbr > 0) & (df.pbr <= 1.0) & (df.roe >= 15.0)].copy()
        value_data["판단_raw"] = value_data["판단"]
        for j, col in enumerate(cols, 1):
            cell = ws2.cell(row=1, column=j, value=col)
            cell.fill = header_fill
            cell.font = header_font
        for i, row in enumerate(value_data.itertuples(), 2):
            ws2.cell(row=i, column=1, value=row.ticker)
            ws2.cell(row=i, column=2, value=row.name)
            ws2.cell(row=i, column=3, value=row.market)
            ws2.cell(row=i, column=4, value=row.pbr)
            ws2.cell(row=i, column=5, value=row.roe)
            ws2.cell(row=i, column=6, value=row.per)
            ws2.cell(row=i, column=7, value=row.eps)
            ws2.cell(row=i, column=8, value=row.marcap)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    df_excel = df.copy()
    df_excel["판단_raw"] = df_excel["판단"]
    excel_bytes = create_screening_excel(df_excel)
    today_str = date.today().strftime("%Y%m%d")
    st.download_button(
        label="📥 엑셀 다운로드",
        data=excel_bytes,
        file_name=f"screening_result_{today_str}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
