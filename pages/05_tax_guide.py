"""절세 전략 가이드 페이지.

ISA · 연금저축 · IRP 3대 절세계좌 비교 및 세후 수익률 시뮬레이션
"""

import io
import sys
from pathlib import Path
from datetime import date

import pandas as pd
import numpy as np
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, str(Path(__file__).parent.parent))

from utils.styles import inject_css, COLOR
from utils.charts import apply_layout

st.set_page_config(page_title="절세 전략", page_icon="💰", layout="wide")
inject_css()

st.title("💰 절세 투자 전략")
st.caption("ISA · 연금저축 · IRP 3대 절세계좌 완전 비교")

# ──────────────────────────────────────────────
# 3대 절세계좌 비교 카드
# ──────────────────────────────────────────────
st.subheader("📋 3대 절세계좌 한눈에 비교")

col1, col2, col3 = st.columns(3)

card_style = lambda color: f"""
  border-top: 4px solid {color}; border-radius: 8px;
  padding: 16px; background: white;
  box-shadow: 0 1px 4px rgba(0,0,0,0.08);
"""

with col1:
    st.markdown(f"""
    <div style="{card_style(COLOR['primary'])}">
      <h3 style="color:{COLOR['primary']}; margin:0 0 12px;">ISA</h3>
      <table style="width:100%; font-size:0.87rem; border-collapse:collapse;">
        <tr><td style="color:#6b7280; padding:4px 0;">연간 한도</td><td><b>2,000만원</b></td></tr>
        <tr><td style="color:#6b7280; padding:4px 0;">세제 혜택</td><td><b>3년 후 9.9% 분리과세</b></td></tr>
        <tr><td style="color:#6b7280; padding:4px 0;">비과세 한도</td><td><b>200만원 (서민형 400만원)</b></td></tr>
        <tr><td style="color:#6b7280; padding:4px 0;">인출</td><td><b>3년 후 자유 인출</b></td></tr>
        <tr><td style="color:#6b7280; padding:4px 0;">투자 대상</td><td><b>국내 주식·ETF·펀드 전체</b></td></tr>
        <tr><td style="color:#6b7280; padding:4px 0;">추천 종목</td><td><b>국내 고배당주, ETF</b></td></tr>
      </table>
    </div>
    """, unsafe_allow_html=True)

with col2:
    st.markdown(f"""
    <div style="{card_style(COLOR['positive'])}">
      <h3 style="color:{COLOR['positive']}; margin:0 0 12px;">연금저축</h3>
      <table style="width:100%; font-size:0.87rem; border-collapse:collapse;">
        <tr><td style="color:#6b7280; padding:4px 0;">연간 한도</td><td><b>1,800만원</b></td></tr>
        <tr><td style="color:#6b7280; padding:4px 0;">세액공제</td><td><b>600만원 × 13.2%~16.5%</b></td></tr>
        <tr><td style="color:#6b7280; padding:4px 0;">세금이연</td><td><b>55세까지 과세이연</b></td></tr>
        <tr><td style="color:#6b7280; padding:4px 0;">수령 세율</td><td><b>3.3~5.5% 연금소득세</b></td></tr>
        <tr><td style="color:#6b7280; padding:4px 0;">투자 대상</td><td><b>국내상장 해외ETF 가능</b></td></tr>
        <tr><td style="color:#6b7280; padding:4px 0;">추천 종목</td><td><b>해외ETF, 커버드콜ETF</b></td></tr>
      </table>
    </div>
    """, unsafe_allow_html=True)

with col3:
    st.markdown(f"""
    <div style="{card_style(COLOR['warning'])}">
      <h3 style="color:{COLOR['warning']}; margin:0 0 12px;">IRP</h3>
      <table style="width:100%; font-size:0.87rem; border-collapse:collapse;">
        <tr><td style="color:#6b7280; padding:4px 0;">연간 한도</td><td><b>연금저축 포함 900만원</b></td></tr>
        <tr><td style="color:#6b7280; padding:4px 0;">세액공제</td><td><b>900만원 × 13.2%~16.5%</b></td></tr>
        <tr><td style="color:#6b7280; padding:4px 0;">세금이연</td><td><b>55세까지 과세이연</b></td></tr>
        <tr><td style="color:#6b7280; padding:4px 0;">제약</td><td><b>안전자산 30% 의무 편입</b></td></tr>
        <tr><td style="color:#6b7280; padding:4px 0;">투자 대상</td><td><b>ETF, 펀드 (안전자산 30%)</b></td></tr>
        <tr><td style="color:#6b7280; padding:4px 0;">추천 종목</td><td><b>채권ETF + 주식ETF 혼합</b></td></tr>
      </table>
    </div>
    """, unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

with st.expander("계좌별 절세 효과 비교"):
    st.markdown("""
    | 구분 | 일반계좌 | ISA | 연금저축 |
    |------|----------|-----|---------|
    | 배당소득세 | 15.4% | 9.9% (비과세 초과분) | 과세이연 |
    | 매매차익 | 비과세 (국내주식) | 비과세 | 과세이연 |
    | 해외ETF 분배금 | 15.4% 원천징수 | 9.9% | 과세이연 → 3.3~5.5% |
    | 커버드콜 분배금 | 15.4% | 9.9% | 과세이연 → 3.3~5.5% |

    **핵심**: 연금저축에서 해외ETF 분배금 → 15.4% → 3.3~5.5%로 세금 절감
    """)

st.divider()

# ──────────────────────────────────────────────
# 종목 유형별 계좌 추천
# ──────────────────────────────────────────────
st.subheader("🎯 종목 유형별 최적 계좌 추천")

stock_type = st.selectbox("투자 종목 유형을 선택하세요", [
    "국내 고배당주 (삼성전자우, 금융주 등)",
    "국내 성장주 (코스피/코스닥 일반주)",
    "국내상장 해외ETF (S&P500, 나스닥 등)",
    "커버드콜 ETF (월배당 ETF)",
    "채권 ETF",
])

reco_map = {
    "국내 고배당주 (삼성전자우, 금융주 등)": {
        "추천 계좌": "ISA 또는 연금저축",
        "이유": "배당소득세 15.4% → ISA에서 9.9%로 절감, 연금저축에서 과세이연",
        "전략": "ISA 비과세 한도(200만원) 초과 배당은 연금저축으로 이동",
        "color": COLOR["primary"],
    },
    "국내 성장주 (코스피/코스닥 일반주)": {
        "추천 계좌": "일반계좌",
        "이유": "국내주식 매매차익은 원래 비과세 → 절세계좌 불필요",
        "전략": "세제 혜택 없으므로 일반계좌에서 자유롭게 거래",
        "color": COLOR["text_muted"],
    },
    "국내상장 해외ETF (S&P500, 나스닥 등)": {
        "추천 계좌": "연금저축 (최우선)",
        "이유": "분배금 15.4% 원천징수 → 연금저축에서 3.3~5.5%로 대폭 절감",
        "전략": "연금저축 한도(1,800만원) 최대 활용, 초과분은 IRP",
        "color": COLOR["positive"],
    },
    "커버드콜 ETF (월배당 ETF)": {
        "추천 계좌": "연금저축 (최우선)",
        "이유": "월배당 분배금 15.4% → 연금저축에서 과세이연, 복리 극대화",
        "전략": "배당 재투자 복리 효과 + 과세이연 이중 혜택",
        "color": COLOR["positive"],
    },
    "채권 ETF": {
        "추천 계좌": "IRP 또는 연금저축",
        "이유": "이자소득세 15.4% → 과세이연으로 복리 효과 극대화",
        "전략": "IRP 안전자산 30% 의무 요건 충족에 활용 가능",
        "color": COLOR["warning"],
    },
}

reco = reco_map[stock_type]
st.markdown(f"""
<div style="background:{reco['color']}15; border-left:4px solid {reco['color']};
            padding:16px 20px; border-radius:8px;">
  <h4 style="color:{reco['color']}; margin:0 0 8px;">✅ 추천: {reco['추천 계좌']}</h4>
  <p style="margin:0 0 6px;"><strong>이유:</strong> {reco['이유']}</p>
  <p style="margin:0;"><strong>전략:</strong> {reco['전략']}</p>
</div>
""", unsafe_allow_html=True)

st.divider()

# ──────────────────────────────────────────────
# 세후 수익률 시뮬레이터 (매년 납입 방식)
# ──────────────────────────────────────────────
st.subheader("📊 세후 수익률 시뮬레이터")
st.info(
    "**시뮬레이터 기준**\n\n"
    "- **투자 금액**: 매년 초에 동일 금액 납입 (적립식)\n"
    "- **수익률 구조**: **복리** (매년 수익이 원금에 합산되어 다시 투자됨)\n"
    "- **세후 수령액**: 투자 기간 종료 시점에 **일시 수령할 경우**의 총 세후 금액\n"
    "- **IRP**: 안전자산 30% 의무 편입 → 실효수익률 = 위험자산 70% × 연수익률 + 안전자산 30% × 3%",
    icon="ℹ️",
)

col_inp1, col_inp2, col_inp3 = st.columns(3)
with col_inp1:
    annual_payment = st.number_input(
        "연간 납입금 (만원)",
        min_value=100, max_value=10000,
        value=500, step=100,
        help="매년 초에 납입하는 금액입니다.",
    )
with col_inp2:
    annual_return = st.slider("연간 기대수익률 (%)", 1.0, 30.0, 8.0, step=0.5,
                               help="복리로 매년 적용되는 연간 수익률입니다.")
with col_inp3:
    invest_years = st.slider("투자 기간 (년)", 1, 30, 10,
                              help="납입 기간이자 투자 기간. 종료 시점에 일시 수령 가정.")

payment = float(annual_payment) * 10000  # 만원 → 원
r = annual_return / 100

def _sim_annuity(payment: float, rate: float, years: int,
                 annual_tax: float = 0.0, withdrawal_tax_on_gain: float = 0.0) -> list[float]:
    """매년 초 납입 적립식 시뮬레이션. 연도별 세후 잔액 반환."""
    balance = 0.0
    total_paid = 0.0
    results = []
    for _ in range(years):
        balance += payment
        total_paid += payment
        gain = balance * rate
        balance += gain * (1 - annual_tax)
        gain_total = max(balance - total_paid, 0)
        results.append(balance - gain_total * withdrawal_tax_on_gain)
    return results

# 일반계좌: 배당/이자 15.4% 과세
general = _sim_annuity(payment, r, invest_years, annual_tax=0.154)

# ISA: 과세이연 → 만기 후 순이익 중 200만원 초과분에 9.9%
def _sim_isa(payment, rate, years):
    balance = 0.0
    total_paid = 0.0
    results_interim = []
    for _ in range(years):
        balance += payment
        total_paid += payment
        balance += balance * rate
        gain = max(balance - total_paid, 0)
        tax_free = 200 * 10000
        taxable_gain = max(gain - tax_free, 0)
        results_interim.append(balance - taxable_gain * 0.099)
    return results_interim

isa_list = _sim_isa(payment, r, invest_years)

# 연금저축: 과세이연 → 수령 시 3.3% 연금소득세
pension = _sim_annuity(payment, r, invest_years, withdrawal_tax_on_gain=0.033)

# IRP: 안전자산 30% 의무 → 실효수익률 낮춤, 수령 시 3.3%
irp_rate = r * 0.7 + 0.03 * 0.3
irp = _sim_annuity(payment, irp_rate, invest_years, withdrawal_tax_on_gain=0.033)

years_axis = list(range(1, invest_years + 1))
total_paid_list = [payment * (i + 1) / 10000 for i in range(invest_years)]

# 차트
fig_sim = go.Figure()
sim_data = [
    ("일반계좌", general, COLOR["text_muted"], "dash"),
    ("ISA", isa_list, COLOR["primary"], "solid"),
    ("연금저축", pension, COLOR["positive"], "solid"),
    ("IRP", irp, COLOR["warning"], "dot"),
]
for label, vals, color, dash in sim_data:
    fig_sim.add_trace(go.Scatter(
        x=years_axis, y=[v / 10000 for v in vals],
        mode="lines", name=label,
        line=dict(color=color, width=2, dash=dash),
    ))
# 총 납입액 기준선
fig_sim.add_trace(go.Scatter(
    x=years_axis, y=total_paid_list,
    mode="lines", name="총 납입액",
    line=dict(color="#d1d5db", width=1.5, dash="dot"),
))

apply_layout(fig_sim, "")
fig_sim.update_layout(
    xaxis_title="투자 기간 (년)",
    yaxis_title="세후 자산 (만원)",
    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    height=400,
)
st.plotly_chart(fig_sim, use_container_width=True)

# 최종 비교 표
total_paid_won = payment * invest_years
final_general = general[-1] / 10000
final_isa = isa_list[-1] / 10000
final_pension = pension[-1] / 10000
final_irp = irp[-1] / 10000
total_paid_man = total_paid_won / 10000

st.caption(
    f"📌 연 납입금 **{annual_payment:,}만원** × {invest_years}년 = 총 **{total_paid_man:,.0f}만원** 납입 "
    f"| 연 수익률 **{annual_return}% 복리** | {invest_years}년 후 **일시 수령** 기준"
)
st.markdown(f"""
| 계좌 | {invest_years}년 후 세후 수령액 | 절세 효과 | 납입 대비 배율 |
|------|------------------------|-----------|------|
| 일반계좌 | **{final_general:,.0f}만원** | 기준 | {final_general/total_paid_man:.2f}배 |
| ISA | **{final_isa:,.0f}만원** | **+{final_isa - final_general:,.0f}만원** | {final_isa/total_paid_man:.2f}배 |
| 연금저축 | **{final_pension:,.0f}만원** | **+{final_pension - final_general:,.0f}만원** | {final_pension/total_paid_man:.2f}배 |
| IRP | **{final_irp:,.0f}만원** | **+{final_irp - final_general:,.0f}만원** | {final_irp/total_paid_man:.2f}배 |
""")

# ──────────────────────────────────────────────
# 세액공제 환급액 계산 함수
# ──────────────────────────────────────────────
def _calc_tax_deduction(pension_input: float, irp_input: float, income_man: float) -> dict:
    """연금저축/IRP 세액공제 환급액 계산."""
    rate = 0.165 if income_man <= 5500 else 0.132
    pension_limit = 600 if income_man <= 12000 else 300
    pension_ded = min(pension_input, pension_limit)
    irp_ded = min(irp_input, max(900 - pension_ded, 0))
    total_ded = pension_ded + irp_ded
    return {
        "rate": rate,
        "pension_ded": pension_ded,
        "irp_ded": irp_ded,
        "total_ded": total_ded,
        "pension_refund": pension_ded * 10000 * rate,
        "irp_refund": irp_ded * 10000 * rate,
        "total_refund": total_ded * 10000 * rate,
        "pension_limit": pension_limit,
        "pension_input": pension_input,
        "irp_input": irp_input,
    }


# XLSX 내보내기
def create_tax_excel(payment, r, years, general, isa_list, pension, irp,
                     tax_info: dict | None = None, pension_r: list | None = None):
    wb = Workbook()
    ws = wb.active
    ws.title = "절세 시뮬레이션"

    header_fill = PatternFill(start_color="2F64E3", end_color="2F64E3", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = ["연도", "총납입액 (만원)", "일반계좌 (만원)", "ISA (만원)", "연금저축 (만원)", "IRP (만원)"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i in range(years):
        ws.cell(row=i + 2, column=1, value=i + 1)
        ws.cell(row=i + 2, column=2, value=round(payment * (i + 1) / 10000, 0))
        ws.cell(row=i + 2, column=3, value=round(general[i] / 10000, 0) if i < len(general) else "")
        ws.cell(row=i + 2, column=4, value=round(isa_list[i] / 10000, 0) if i < len(isa_list) else "")
        ws.cell(row=i + 2, column=5, value=round(pension[i] / 10000, 0) if i < len(pension) else "")
        ws.cell(row=i + 2, column=6, value=round(irp[i] / 10000, 0) if i < len(irp) else "")

    # Sheet2: 세액공제 시뮬레이션
    if tax_info and pension_r:
        ws2 = wb.create_sheet("세액공제 시뮬레이션")
        # 요약 정보
        summary = [
            ("총급여 (만원)", tax_info.get("income_man", "-")),
            ("적용 세율", f"{tax_info['rate']*100:.1f}%"),
            ("연금저축 공제 대상액 (만원)", tax_info["pension_ded"]),
            ("IRP 공제 대상액 (만원)", tax_info["irp_ded"]),
            ("연간 세액공제 환급액 (만원)", round(tax_info["total_refund"] / 10000, 1)),
        ]
        for row_i, (k, v) in enumerate(summary, 1):
            ws2.cell(row=row_i, column=1, value=k).font = Font(bold=True)
            ws2.cell(row=row_i, column=2, value=v)

        ws2.cell(row=7, column=1, value="연도").font = Font(bold=True)
        ws2.cell(row=7, column=2, value="연금저축 (만원)").font = Font(bold=True)
        ws2.cell(row=7, column=3, value="세액공제 재투자 포함 (만원)").font = Font(bold=True)
        ws2.cell(row=7, column=4, value="누적 환급액 (만원)").font = Font(bold=True)
        for i in range(years):
            ws2.cell(row=i + 8, column=1, value=i + 1)
            ws2.cell(row=i + 8, column=2, value=round(pension[i] / 10000, 0) if i < len(pension) else "")
            ws2.cell(row=i + 8, column=3, value=round(pension_r[i] / 10000, 0) if i < len(pension_r) else "")
            ws2.cell(row=i + 8, column=4, value=round(tax_info["total_refund"] / 10000 * (i + 1), 1))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────
# 세액공제 환급액 시뮬레이션 섹션
# ──────────────────────────────────────────────
st.divider()
st.subheader("💸 세액공제 환급액 시뮬레이션")
st.caption("연금저축·IRP 납입 시 연말정산에서 돌려받는 세액공제 환급액과 재투자 효과를 계산합니다.")

col_td1, col_td2, col_td3 = st.columns(3)
with col_td1:
    income_man = st.number_input(
        "총급여 (만원)",
        min_value=1000, max_value=100000,
        value=5000, step=100,
        help="연간 총급여액. 5,500만원 이하 → 16.5%, 초과 → 13.2% 세율 적용",
    )
with col_td2:
    pension_input = st.number_input(
        "연금저축 연간 납입액 (만원)",
        min_value=0, max_value=1800,
        value=int(min(annual_payment, 600)),
        step=50,
        help="세액공제 한도: 총급여 1.2억 이하 600만원 / 초과 300만원",
    )
with col_td3:
    irp_input = st.number_input(
        "IRP 연간 납입액 (만원)",
        min_value=0, max_value=900,
        value=0, step=50,
        help="연금저축+IRP 합산 세액공제 한도 900만원",
    )

td = _calc_tax_deduction(pension_input, irp_input, income_man)

# 한도 초과 경고
warnings = []
if pension_input > td["pension_limit"]:
    warnings.append(f"연금저축 납입액({pension_input}만원)이 공제 한도({td['pension_limit']}만원)를 초과합니다. 초과분은 세액공제 미적용.")
if irp_input > td["irp_ded"]:
    warnings.append(f"IRP 납입액({irp_input}만원) 중 일부가 합산 한도(900만원) 초과로 세액공제 미적용.")
for w in warnings:
    st.warning(w)

# 세액공제 요약 카드
rate_pct = td["rate"] * 100
c1, c2, c3 = st.columns(3)
with c1:
    st.markdown(f"""
    <div style="background:{COLOR['positive']}15; border-left:4px solid {COLOR['positive']};
                padding:16px; border-radius:8px; text-align:center;">
      <div style="font-size:0.85rem; color:#6b7280;">연금저축 환급액</div>
      <div style="font-size:1.6rem; font-weight:700; color:{COLOR['positive']};">
        {td['pension_refund']/10000:,.0f}만원
      </div>
      <div style="font-size:0.8rem; color:#6b7280;">
        {td['pension_ded']}만원 × {rate_pct:.1f}%
      </div>
    </div>
    """, unsafe_allow_html=True)
with c2:
    st.markdown(f"""
    <div style="background:{COLOR['warning']}15; border-left:4px solid {COLOR['warning']};
                padding:16px; border-radius:8px; text-align:center;">
      <div style="font-size:0.85rem; color:#6b7280;">IRP 환급액</div>
      <div style="font-size:1.6rem; font-weight:700; color:{COLOR['warning']};">
        {td['irp_refund']/10000:,.0f}만원
      </div>
      <div style="font-size:0.8rem; color:#6b7280;">
        {td['irp_ded']}만원 × {rate_pct:.1f}%
      </div>
    </div>
    """, unsafe_allow_html=True)
with c3:
    st.markdown(f"""
    <div style="background:{COLOR['primary']}15; border-left:4px solid {COLOR['primary']};
                padding:16px; border-radius:8px; text-align:center;">
      <div style="font-size:0.85rem; color:#6b7280;">연간 합계 환급액</div>
      <div style="font-size:1.6rem; font-weight:700; color:{COLOR['primary']};">
        {td['total_refund']/10000:,.0f}만원
      </div>
      <div style="font-size:0.8rem; color:#6b7280;">
        {invest_years}년 누적 {td['total_refund']/10000*invest_years:,.0f}만원
      </div>
    </div>
    """, unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ── 기존 시뮬레이터와 연동: 재투자 효과 ──
annual_refund = td["total_refund"]  # 원 단위
pension_r = _sim_annuity(payment + annual_refund, r, invest_years, withdrawal_tax_on_gain=0.033)
final_pension_r = pension_r[-1] / 10000
total_refund_cum = annual_refund / 10000 * invest_years
extra_gain = final_pension_r - final_pension

st.markdown("#### 세액공제 환급액 재투자 시 실질 수령액")
st.caption(f"연간 환급액 **{annual_refund/10000:,.0f}만원**을 매년 연금저축에 추가 납입·복리 운용 가정 | {invest_years}년 기준")

st.markdown(f"""
| 항목 | 세액공제 미반영 | 세액공제 재투자 포함 | 차이 |
|------|----------------|---------------------|------|
| 연금저축 수령액 | **{final_pension:,.0f}만원** | **{final_pension_r:,.0f}만원** | **+{extra_gain:,.0f}만원** |
| 세액공제 누적 환급액 | — | {total_refund_cum:,.0f}만원 납입 추가 | |
| 납입 대비 배율 | {final_pension/total_paid_man:.2f}배 | {final_pension_r/(total_paid_man + total_refund_cum):.2f}배 | |
""")

# XLSX 내보내기 (Sheet2 포함)
tax_info_for_excel = {**td, "income_man": income_man}
today_str = date.today().strftime("%Y%m%d")
excel_bytes = create_tax_excel(
    payment, r, invest_years, general, isa_list, pension, irp,
    tax_info=tax_info_for_excel, pension_r=pension_r,
)
st.download_button(
    label="📥 시뮬레이션 저장 (Excel)",
    data=excel_bytes,
    file_name=f"tax_simulation_{today_str}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
